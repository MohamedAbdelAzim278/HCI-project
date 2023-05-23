import os
import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
from openpyxl import load_workbook

def add_user():
    global name_entry, age_entry, photo_path, photo_label

    # Get the name and age from the entry fields
    name = name_entry.get()
    age = age_entry.get()

    # Load the Excel sheet
    workbook = load_workbook("Users.xlsx")
    sheet = workbook["Sheet1"]  # Update with the correct sheet name

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the user data to the new row
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = age
    #sheet.cell(row=next_row, column=3).value = photo_path

    # Save the updated Excel sheet
    workbook.save("Users.xlsx")

    # Clear the entry fields
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    photo_path = ""
    update_name_listbox()
    # Clear the photo label
    photo_label.config(image=None)
def delete_user():
    global name_entry, name_listbox

    # Get the selected name from the listbox
    selected_name = name_listbox.get(tk.ACTIVE)

    # Load the Excel sheet
    workbook = load_workbook("Users.xlsx")
    sheet = workbook["Sheet1"]  # Update with the correct sheet name

    # Find the row that matches the selected name
    row_index = None
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == selected_name:
            row_index = index
            break

    if row_index is not None:
        # Delete the row
        sheet.delete_rows(row_index)

        # Save the updated Excel sheet
        workbook.save("Users.xlsx")

    # Clear the entry field
    name_entry.delete(0, tk.END)

    # Update the listbox with the updated names
    update_name_listbox()


def update_name_listbox():
    global name_listbox

    # Load the Excel sheet
    workbook = load_workbook("Users.xlsx")
    sheet = workbook["Sheet1"]  # Update with the correct sheet name

    # Clear the listbox
    name_listbox.delete(0, tk.END)

    # Get all the names from the Excel sheet
    names = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Add the names to the listbox
    for name in names:
        name_listbox.insert(tk.END, name)


def upload_photo():
    global photo_path, photo_label

    # Open a file dialog to select a photo
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png")])

    # Update the photo path variable
    photo_path = file_path

    # Display the photo on the label
    image = Image.open(photo_path)
    image = image.resize((500, 300))  # Adjust the size as desired
    photo = ImageTk.PhotoImage(image)
    photo_label.config(image=photo)
    photo_label.image = photo

    # Save the photo in the "photos" folder with the name entered by the user
    save_folder = "photos"
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    name = name_entry.get()
    file_name = f"{name}.jpg"  # You can change the extension as needed
    save_path = os.path.join(save_folder, file_name)
    image.save(save_path)

def create_interface():
    global name_entry, age_entry, photo_path, photo_label, name_listbox

    # Create the main window
    window = tk.Tk()
    window.title("User Registration")

    # Set the size of the main window
    window.geometry("400x400")  # Set the width and height as desired

    # Create a label and entry field for the name
    name_label = tk.Label(window, text="Name:")
    name_label.pack()
    name_entry = tk.Entry(window)
    name_entry.pack()

    # Create a label and entry field for the age
    age_label = tk.Label(window, text="Age:")
    age_label.pack()
    age_entry = tk.Entry(window)
    age_entry.pack()

    # Create a button to upload a photo
    upload_button = tk.Button(window, text="Upload Photo", command=upload_photo)
    upload_button.pack()

    # Create a label to display the selected photo
    photo_label = tk.Label(window)
    photo_label.pack()

    # Create a listbox to display the names
    name_listbox = tk.Listbox(window)
    name_listbox.pack()

    # Update the listbox with the names
    update_name_listbox()

    # Create a button to add the user
    add_button = tk.Button(window, text="Add User", command=add_user)
    add_button.pack()

    # Create a button to delete the user
    delete_button = tk.Button(window, text="Delete User", command=delete_user)
    delete_button.pack()

    # Start the main event loop
    window.mainloop()

# Call the function to create the interface
photo_path = ""  # Variable to store the selected photo path
create_interface()